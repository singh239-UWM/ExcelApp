# libs
from tkinter import *
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk

from openpyxl.descriptors.base import String
#from openpyxl import Workbook

# globla varible
filepath = ""

#Function for opening the file explorer window
def browseFiles():
    global filepath
    filepath = filedialog.askopenfilename(initialdir = "/", 
                                          title = "Select a File", 
                                          filetypes = (("Microsoft Excel Worksheet", "*.xlsx*"), ("All Files", "*.*")) )
    #Change fileLable contents
    fileLable.configure(text="File Opened: "+ filepath)


def generateData():
    table.delete(*table.get_children())
    # print(filepath)
    if filepath == "":
        print("Select File")
    elif ".xlsx" not in filepath:
        print("Select .xlsx file")
    else:
        wb = load_workbook(filepath) #open workbook
        ws = wb.active #active first worksheet
        # c = ws['A4']
        # colE = ws['E']
        # colF = ws['F']
        # print(c.value)
        # print(colE)
        # for cell in colF:
        #     print(cell.value)
        colRange = len(ws[1])
        rowRange = len(ws['A'])
        # print(colRange, " ", rowRange)
        # print(colRange[1].value)
        # for row in ws.iter_rows(min_row = 1, max_col = colRange, max_row = rowRange):
        #     for cell in row:
        #         print(cell.value, end = " ")
        #     print()
        headingTuple = ()

        #Making tuple for heading column
        for i in range(colRange):
            temp = 'A'
            tempTuple = (chr(ord(temp) + i),)
            headingTuple = headingTuple + tempTuple
        
        #tuple of colums id for table
        table['columns'] = headingTuple
        
        #nitializing heading columns
        for i in range(colRange+1):
            if i == 0:
                table.column('#0', width = 40, stretch=NO) #need for impo
            else:
                table.column(headingTuple[i-1], width = 80, anchor=CENTER)

        # displaying heading 
        for i in range(colRange+1):
            if i == 0:
                table.heading('#0', text='', anchor=CENTER)
            else:
                table.heading(headingTuple[i-1], text=headingTuple[i-1], anchor=CENTER)

        # insering data
        # table.insert(parent='', index=0, text='1', values=headingTuple)
        for i in range(rowRange):   
            cols = ws[i+1]
            tempColsTuple = ()
            for j in range(colRange):
                #create tuble
                tempVal = cols[j].value
                tempTuple = (tempVal,)
                tempColsTuple = tempColsTuple + tempTuple
            # insert data on table
            table.insert(parent='', index=i, text=i, values=tempColsTuple)

# create window object and set its properties
window = tk.Tk()    #main window
window.title("File") #title name
window.geometry('1250x900') #size of the displayed window
window.pack_propagate(False) # widger will not resize windoe
window.resizable(0,0) # user will not resize window
# window.config(background="black")

#frame for opening file
fileFrame = tk.LabelFrame(window, text="Open File")
fileFrame.place(height=110, width=1200, rely=0.01, relx=0.01)

#frame for showing excel data
exclFrame = tk.LabelFrame(window, text="Excel Data")
exclFrame.place(height=735, width=1200, rely=0.15, relx=0.01)


  
#browse file button
button_explore = tk.Button(fileFrame, text = "Browse Files", background = "green", command = browseFiles)
button_explore.place(relx = 0.90, rely=0)
#exit button
button_exit = tk.Button(fileFrame, text = "Exit", background = "red", height = 1)
button_exit.place(relx = 0.90, rely=0.333)
#generate button
button_generate = tk.Button(fileFrame, text = "Generate Table", command = generateData)
button_generate.place(relx = 0.90, rely=0.666)
# Create a File name label
fileLable = tk.Label(fileFrame, text = "No File Selected", fg = "blue")
fileLable.place(relx=0.01, rely=0.3)
# treeview table
table = ttk.Treeview(exclFrame)
table.place(relheight=1, relwidth=1)
# scrollbar for treeview table
tableScrollY = tk.Scrollbar(exclFrame, orient="vertical", command=table.yview)
tableScrollX = tk.Scrollbar(exclFrame, orient="horizontal", command=table.xview)
# configs for scroll bar
table.config(xscrollcommand=tableScrollX.set, yscrollcommand=tableScrollY.set) #tell table to take this scrollbar
tableScrollX.pack(side="bottom", fill="x")
tableScrollY.pack(side="right", fill="y")
#





# tab.pack(fill = "both")

# Grid method is chosen for placing
# the widgets at respective positions
# in a table like structure by
# specifying rows and columns
# label_file_explorer.grid(column = 1, row = 1, padx=20)
# table.grid(column = 1, row = 4, pady = 10)
# Start program
window.mainloop()
